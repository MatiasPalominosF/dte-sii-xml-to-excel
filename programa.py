import pandas as pd
import xml.etree.ElementTree as ET
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Función para parsear un archivo XML y extraer los datos en forma de lista de diccionarios
def parse_xml_to_dict(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    data = []
    # Recorrer todos los elementos 'DTE' en el XML
    for dte in root.findall('DTE'):
        # Recorrer todos los elementos 'Documento' en 'DTE'
        for doc in dte.findall('Documento'):
            doc_data = {}

            # Procesar la sección 'Encabezado'
            encabezado = doc.find('Encabezado')
            if encabezado is not None:
                # Extraer información de 'IdDoc'
                iddoc = encabezado.find('IdDoc')
                if iddoc is not None:
                    doc_data['TipoDTE'] = iddoc.findtext('TipoDTE')
                    doc_data['Folio'] = iddoc.findtext('Folio')
                    doc_data['FchEmis'] = iddoc.findtext('FchEmis')
                    doc_data['IndTraslado'] = iddoc.findtext('IndTraslado')

                # Extraer información de 'Emisor'
                emisor = encabezado.find('Emisor')
                if emisor is not None:
                    doc_data['RUTEmisor'] = emisor.findtext('RUTEmisor')
                    doc_data['RznSocEmisor'] = emisor.findtext('RznSoc')
                    doc_data['GiroEmis'] = emisor.findtext('GiroEmis')
                    doc_data['TelefonoEmisor'] = emisor.findtext('Telefono')
                    doc_data['CorreoEmisor'] = emisor.findtext('CorreoEmisor')
                    doc_data['Acteco'] = emisor.findtext('Acteco')
                    doc_data['CdgSIISucur'] = emisor.findtext('CdgSIISucur')
                    doc_data['DirOrigen'] = emisor.findtext('DirOrigen')
                    doc_data['CmnaOrigen'] = emisor.findtext('CmnaOrigen')
                    doc_data['CiudadOrigen'] = emisor.findtext('CiudadOrigen')

                # Extraer información de 'Receptor'
                receptor = encabezado.find('Receptor')
                if receptor is not None:
                    doc_data['RUTRecep'] = receptor.findtext('RUTRecep')
                    doc_data['RznSocRecep'] = receptor.findtext('RznSocRecep')
                    doc_data['GiroRecep'] = receptor.findtext('GiroRecep')
                    doc_data['DirRecep'] = receptor.findtext('DirRecep')
                    doc_data['CmnaRecep'] = receptor.findtext('CmnaRecep')
                    doc_data['CiudadRecep'] = receptor.findtext('CiudadRecep')

                # Extraer información de 'Totales'
                totales = encabezado.find('Totales')
                if totales is not None:
                    doc_data['MntNeto'] = totales.findtext('MntNeto')
                    doc_data['TasaIVA'] = totales.findtext('TasaIVA')
                    doc_data['IVA'] = totales.findtext('IVA')
                    doc_data['MntTotal'] = totales.findtext('MntTotal')

            # Extraer información de 'Detalle'
            detalle = doc.find('Detalle')
            if detalle is not None:
                doc_data['NroLinDet'] = detalle.findtext('NroLinDet')
                doc_data['NmbItem'] = detalle.findtext('NmbItem')
                doc_data['QtyItem'] = detalle.findtext('QtyItem')
                doc_data['UnmdItem'] = detalle.findtext('UnmdItem')
                doc_data['PrcItem'] = detalle.findtext('PrcItem')
                doc_data['MontoItem'] = detalle.findtext('MontoItem')

            # Extraer el valor de 'FRMA'
            ted = doc.find('TED')
            if ted is not None:
                dd = ted.find('DD')
                if dd is not None:
                    caf = dd.find('CAF')
                    if caf is not None:
                        doc_data['FIRMA'] = caf.findtext('FRMA')
                        da = caf.find('DA')
                        if da is not None:
                            doc_data['FECHA ACEPTACIÓN'] = da.findtext('FA')

            # Añadir el diccionario de datos a la lista de datos
            data.append(doc_data)

    return data

def adjust_column_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
        
# Directorio que contiene los archivos XML
xml_directory = 'archivos/'

# Lista para almacenar todos los datos
all_data = []

# Recorrer todos los archivos en el directorio
for filename in os.listdir(xml_directory):
    if filename.endswith('.xml'):
        xml_file = os.path.join(xml_directory, filename)
        data = parse_xml_to_dict(xml_file)
        all_data.extend(data)

# Crear un DataFrame de pandas con todos los datos
df = pd.DataFrame(all_data)

# Guardar el DataFrame en un archivo Excel
excel_file = 'DTE_Recibidos_Combined.xlsx'
df.to_excel(excel_file, index=False)

# Abrir el archivo Excel y ajustar las columnas y filas
wb = load_workbook(excel_file)
ws = wb.active

# Ajustar el ancho de las columnas
adjust_column_width(ws)

# Ajustar la altura de las filas
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].auto_size = True

# Guardar los cambios
wb.save(excel_file)

print(f'Datos guardados en {excel_file} con columnas y filas ajustadas')
